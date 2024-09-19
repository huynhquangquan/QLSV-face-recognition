import openpyxl
from openpyxl.workbook import workbook
from openpyxl import load_workbook
from datetime import datetime
import numbers


# Mở file Excel/Tạo Excel
def diary_attandance(username,id):
    try:
        wb = load_workbook('attendance.xlsx')
        # Chọn sheet mặc định
        check_row=False
        sheet = wb.active
        current = sheet.max_row
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        # Đọc dữ liệu từ cột 'Name'
        print("Before")
        for row in sheet.iter_rows(values_only=True):
            print(row[1])  # In ra tên sinh viên
            if row[1]:
                check_row=True
                print(current)
                break
            else:
                print(current-1)
                break
        # print(check_row)
        # Cập nhật dữ liệu
        if check_row:
            sheet[f'A{current + 1}'] = f'{id}'
            sheet[f'B{current + 1}'] = f'{username}'
            sheet[f'C{current + 1}'] = datetime.now().strftime('%d-%m-%Y')
            sheet[f'D{current + 1}'] = datetime.now().strftime('%H:%M')
        else:
            sheet[f'A{current}'] = f'{id}'
            sheet[f'B{current}'] = f'{username}'
            sheet[f'C{current}'] = datetime.now().strftime('%d-%m-%Y')
            sheet[f'D{current}'] = datetime.now().strftime('%H:%M')

        # Lưu thay đổi
        wb.save('attendance.xlsx')

        # Đọc dữ liệu từ cột 'Name'
        print("After")
        for row in sheet.iter_rows(values_only=True):
            print(row[1])  # In ra tên sinh viên
        print(sheet.max_row)
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 10
        sheet[f'A1'] = f'{id}'
        sheet[f'B1'] = f'{username}'
        sheet[f'C1'] = datetime.now().strftime('%d-%m-%Y')
        sheet[f'D1'] = datetime.now().strftime('%H:%M')
        wb.save('attendance.xlsx')