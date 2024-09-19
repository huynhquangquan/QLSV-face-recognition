import subprocess
from tkinter import *
from tkinter import ttk,messagebox,Button,Frame,Tk, PhotoImage, filedialog
from tkinter.filedialog import asksaveasfilename
from PIL import Image, ImageTk
import tkinter as tk
import managing_function
import mysql.connector
import os
import time
from threading import Thread
import cv2
import customtkinter
from datetime import datetime
import excel_attandance
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from os import sys
import shutil

global verify
verify = False
global username
global id
password = ''
username = ''
global admin_window
admin_window = None
global value
value = None
global id
id = ''
global idCoVan
idCoVan = ''
global idMaBangDiem
idMaBangDiem = ''

while True:
    def main():
        # Default Window
        global  control_user_logout
        control_user_logout = 0
        def on_close():
            result = messagebox.askyesno("Exit", "Bạn muốn thoát ứng dụng?")
            if result:
                verify = False
                sys.exit(0)
        window = customtkinter.CTk()
        window.geometry("950x600+350+120")
        window.title("Ứng dụng quản lý sinh viên")
        window.resizable(False,False)
        icon = PhotoImage(file='resources/logo.png')
        window.iconphoto(True,icon)
        window.config(background="#393939")
        mainpanel_logo = PhotoImage(file='resources/user.png')


        # Frame (Right)
        mainpage_right_panel = Frame(window,bg="#1f1f1f",bd=5, relief=RAISED,width=937,height=680)
        mainpage_right_panel.grid(pady=25, row=0,column=1, rowspan=3, columnspan=4)
        mainpage_right_panel.grid_propagate(0)

        # Frame (Left)
        if username == 'admin' and password == 'admin':
            mainpage_left_panel = Frame(window, bg="#1f1f1f", bd=5, relief=RAISED, pady=238, padx=18)
        else:
            mainpage_left_panel = Frame(window, bg="#1f1f1f", bd=5, relief=RAISED, pady=238, padx=60)
        mainpage_left_panel.grid(row=0,column=0,rowspan=3)

        label_option_0 = Label(mainpage_left_panel, text="TRANG CHỦ", font=('Arial', 15, 'underline'), fg='gray', bg="#1f1f1f")
        label_option_0.grid(row=1, pady=5)
        label_option_0.bind("<Button-1>", lambda event: change_to_main_screen())

        label_option_1 = Label(mainpage_left_panel, text="NHẬT KÝ ĐIỂM DANH", font=('Arial', 15, 'underline'), fg='gray', bg="#1f1f1f")
        label_option_1.bind("<Button-1>", lambda event: change_to_diary())

        label_option_2 = Label(mainpage_left_panel, text="CHI TIẾT", font=('Arial', 15, 'underline'), fg='gray', bg="#1f1f1f")
        label_option_2.bind("<Button-1>", lambda event: change_to_detail())

        label_option_3 = Label(mainpage_left_panel, text="BẢNG ĐIỂM", font=('Arial', 15, 'underline'), fg='gray', bg="#1f1f1f")
        label_option_3.bind("<Button-1>", lambda event: change_to_score())

        label_option_4 = Label(mainpage_left_panel, text="ADMIN/SETTING", font=('Arial', 15, 'underline'), fg='gray',bg="#1f1f1f")
        if username == 'admin' and password == 'admin':
            mainpage_left_panel = Frame(window, bg="#1f1f1f", bd=5, relief=RAISED, pady=238, padx=18)
            label_option_1.grid(row=2, pady=5)
            label_option_4.grid(row=5, pady=5)
        else:
            mainpage_left_panel = Frame(window, bg="#1f1f1f", bd=5, relief=RAISED, pady=238, padx=60)
            label_option_2.grid(row=3, pady=5)
            label_option_3.grid(row=4, pady=5)
        label_option_4.bind("<Button-1>", lambda event: change_to_setting())

        label_username = Label(window, text=f"{username}", font=('Arial', 15, 'underline'), fg='gray',bg="#1f1f1f")
        label_username.grid(row=0, pady=(0,20))
        label_username.bind("<Button-1>", lambda event: on_click_user_option())
        logout = Label(window, text="Đăng Xuất", font=('Arial', 15, 'underline'), fg='gray', bg="#1f1f1f")
        logout.bind("<Button-1>", lambda event: logout_to_login())


        # Function
        def read_excel_data():
            try:
                wb = load_workbook('attendance.xlsx')
                sheet = wb.active
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append({'ID': row[0], 'Họ và Tên': row[1], 'D-M-Y': row[2], 'Thời Gian Điểm Danh': row[3]})
                return data
            except FileNotFoundError:
                return []

        def logout_to_login():
            global id
            global verify
            verify = False
            username = ''
            id = None
            window.quit()
            window.destroy()

        # User's option
        def on_click_user_option():
            global control_user_logout
            if control_user_logout == 1:
                control_user_logout = 0
            else: control_user_logout = 1
            show_user_option()

        def show_user_option():
            global control_user_logout
            if control_user_logout == 1:
                logout.place(relx=0.108,rely=0.225,anchor=CENTER)
            else: logout.place_forget()

        # Diary's panel
        def change_to_diary():
            for widget in mainpage_right_panel.winfo_children():
                widget.destroy()

            def clear_attendance_file():
                if messagebox.askyesno("Clear History", "Bạn muốn xóa lịch sử điểm danh ?"):
                    folder_name = 'attendance.xlsx'
                    wb = load_workbook(folder_name)
                    sheet = wb.active
                    sheet.delete_rows(1,sheet.max_row)
                    wb.save(folder_name)
                    global data
                    # data.clear()
                    data = read_excel_data()
                    # print(type(data))
                    change_to_diary()

            def save_history():
                # import jpype # hàng premium, phải mua để xóa watermark
                # import asposecells
                # jpype.startJVM()
                # from asposecells.api import Workbook

                # workbook = Workbook("attendance.xlsx")

                wb = load_workbook('attendance.xlsx')

                save_path=filedialog.asksaveasfilename(filetypes=[('Excel file','*.xlsx')],
                                                  defaultextension='*.xlsx',
                                                  title="Export History",
                                                  initialdir="/")

                wb.save(save_path)
                # workbook.save(save_path)
                # jpype.shutdownJVM()

            detail_label = Label(mainpage_right_panel, text="Nhật Ký Điểm Danh Sinh Viên",
                                 font=('Arial', 20, 'bold'), fg='#CFE2F3',bg="#1f1f1f")
            detail_label.grid(row=0, column=0, padx=10)

            table_frame = Frame(mainpage_right_panel)
            table_frame.place(relx=0.5, rely=0.5, anchor=CENTER, width=900, height=550)

            y_scroll = Scrollbar(table_frame, orient=VERTICAL)
            x_scroll = Scrollbar(table_frame, orient=HORIZONTAL)

            style = ttk.Style(mainpage_right_panel)

            style.theme_use("clam")

            style.configure("Treeview", background="#1f1f1f", fieldbackground="#1f1f1f", foreground="white", font=10)
            style.configure("Treeview.Heading", background="#393939", foreground="white", font=15)
            style.map("Treeview.Heading", background=[('active', '#393939',)], foreground=[('active', 'white')])
            style.configure('Vertical.TScrollbar', background="#1f1f1f", troughcolor="#393939",
                            gripcount=0, gripwidth=15,gripperlength=20)
            style.configure('Horizontal.TScrollbar', background="#1f1f1f", troughcolor="#393939",
                            gripcount=0, gripwidth=15,gripperlength=20)

            detail_tree = ttk.Treeview(table_frame, yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set,style="Treeview")

            y_scroll.config(command=detail_tree.yview)
            x_scroll.config(command=detail_tree.xview)

            y_scroll.pack(side=RIGHT, fill=Y)
            x_scroll.pack(side=BOTTOM, fill=X)

            detail_tree['columns'] = ("ID", "Họ và Tên","D-M-Y", "Thời Gian Điểm Danh")

            detail_tree.column("ID", anchor=CENTER, width=150, minwidth=150)
            detail_tree.column("Họ và Tên", anchor=CENTER, width=200, minwidth=200)
            detail_tree.column("D-M-Y", anchor=CENTER, width=100, minwidth=100)
            detail_tree.column("Thời Gian Điểm Danh", anchor=CENTER, width=150, minwidth=150)

            detail_tree.heading("ID", anchor=CENTER, text="ID")
            detail_tree.heading("Họ và Tên", anchor=CENTER, text="Họ và Tên")
            detail_tree.heading("D-M-Y", anchor=CENTER, text="D-M-Y")
            detail_tree.heading("Thời Gian Điểm Danh", anchor=CENTER, text="Thời Gian Điểm Danh")

            detail_tree['show'] = 'headings'
            detail_tree.pack(fill=BOTH, expand=1)
            for item in data:
                detail_tree.insert('', 'end', values=(item['ID'], item['Họ và Tên'], item['D-M-Y'], item['Thời Gian Điểm Danh']))

            clear_excel = customtkinter.CTkButton(mainpage_right_panel,text="Clear History",width=100,fg_color="#393939",command=clear_attendance_file)
            clear_excel.grid(row=0,column=1,padx=(120,0), pady=10)
            export_excel = customtkinter.CTkButton(mainpage_right_panel, text="Export History", width=100,fg_color="#393939",command=save_history)
            export_excel.grid(row=0, column=2, padx=(50, 0), pady=10)

        # Detail's Panel
        def change_to_detail():
            for widget in mainpage_right_panel.winfo_children():
                widget.destroy()

            detail_label = Label(mainpage_right_panel,text="THÔNG TIN CHI TIẾT",font=('Arial',20,'bold'),fg='#CFE2F3',bg="#1f1f1f")
            detail_label.grid(row=0,padx=330)
            # connect db
            db = mysql.connector.connect(user='root', password='', host='localhost', database='qlthongtinsv')
            cursor = db.cursor()

            #####

            # top.resizable(False, False) # khong dc phong to cua so
            # id = 3

            def getCurrentStudent(id):
                editCursor = db.cursor()
                sql = "SELECT * FROM SinhVien WHERE idSinhVien = %s"
                data = (id,)

                editCursor.execute(sql, data)
                result = editCursor.fetchone()
                return result

            curentstudent = getCurrentStudent(id)
            print(curentstudent)

            ####

            # idCoVan = 2

            def getCurrentStudent(idCoVan):
                editCursor = db.cursor()
                sql = "SELECT * FROM nguoicovan WHERE macv = %s"
                data = (idCoVan,)

                editCursor.execute(sql, data)
                result = editCursor.fetchone()
                return result

            curentcovan = getCurrentStudent(idCoVan)
            print(curentcovan)
            ####

            # idMaBangDiem = 1

            def getCurrentStudent(idMaBangDiem):
                editCursor = db.cursor()
                sql = "SELECT * FROM bangdiemchitiet WHERE mabangdiem = %s"
                data = (idMaBangDiem,)

                editCursor.execute(sql, data)
                result = editCursor.fetchone()
                return result

            curentBangDiem = getCurrentStudent(idMaBangDiem)
            print(curentBangDiem)
            ####

            customtkinter.CTkLabel(master=mainpage_right_panel, text="Thông tin sinh viên", text_color="red",
                                   font=('Times New Roman', 20, 'bold')).place(relx=0.2, rely=0.05)
            # name
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Tên sinh viên").place(relx=0.1, rely=0.1)
            inputName = customtkinter.CTkLabel(master=mainpage_right_panel, width=200)
            inputName.configure(text=curentstudent[1])
            inputName.place(relx=0.3, rely=0.1)

            # date
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Ngày sinh").place(relx=0.1, rely=0.15)
            inputDate = customtkinter.CTkLabel(master=mainpage_right_panel, width=180)
            inputDate.configure(text=curentstudent[2])
            inputDate.place(relx=0.3, rely=0.15)

            # gender

            # cccd
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Số cccd").place(relx=0.1, rely=0.2)
            inputCccd = customtkinter.CTkLabel(master=mainpage_right_panel, width=180)
            inputCccd.configure(text=curentstudent[5])
            inputCccd.place(relx=0.3, rely=0.2)

            # phone Number
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Số điện thoại").place(relx=0.1, rely=0.25)
            inputPhone = customtkinter.CTkLabel(master=mainpage_right_panel, width=180)
            inputPhone.configure(text=curentstudent[4])
            inputPhone.place(relx=0.3, rely=0.25)

            # email
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Email").place(relx=0.1, rely=0.3)
            inputEmail = customtkinter.CTkLabel(master=mainpage_right_panel, width=180)
            inputEmail.configure(text=curentstudent[6])
            inputEmail.place(relx=0.3, rely=0.3)

            # Place Born
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Nơi sinh").place(relx=0.1, rely=0.35)
            inputBorn = customtkinter.CTkLabel(master=mainpage_right_panel, width=180)
            inputBorn.configure(text=curentstudent[7])
            inputBorn.place(relx=0.3, rely=0.35)

            # nation
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Dân tộc").place(relx=0.1, rely=0.4)
            inputNation = customtkinter.CTkLabel(master=mainpage_right_panel, width=180)
            inputNation.configure(text=curentstudent[8])
            inputNation.place(relx=0.3, rely=0.4)

            # religion
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Tôn giáo").place(relx=0.1, rely=0.45)
            inputReligion = customtkinter.CTkLabel(master=mainpage_right_panel, width=180)
            inputReligion.configure(text=curentstudent[9])
            inputReligion.place(relx=0.3, rely=0.45)

            # Household
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Hộ khẩu").place(relx=0.1, rely=0.50)
            inputHousehold = customtkinter.CTkLabel(master=mainpage_right_panel, width=200)
            inputHousehold.configure(text=curentstudent[10])
            inputHousehold.place(relx=0.3, rely=0.50)

            # adsider
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Tên người cố vấn").place(relx=0.1, rely=0.55)
            inputAdsiderCode = customtkinter.CTkLabel(master=mainpage_right_panel, width=160)
            inputAdsiderCode.configure(text=curentcovan[1])
            inputAdsiderCode.place(relx=0.3, rely=0.55)

            # score
            customtkinter.CTkLabel(master=mainpage_right_panel, text="Điểm trung bình").place(relx=0.1, rely=0.60)
            inputScoreCode = customtkinter.CTkLabel(master=mainpage_right_panel, width=160)
            inputScoreCode.configure(text=(curentBangDiem[3] + curentBangDiem[4]) / 2)
            inputScoreCode.place(relx=0.25, rely=0.60)

            # image
            ###

            # lb = tkinter.Label(top, image=img, text='Ảnh', compound='center')
            ###
            input_image = None

            def chenAnh(id):
                connection, cursor = managing_function.connect_to_database()
                cursor.execute("SELECT url FROM khoanh WHERE url LIKE %s", ('%'+str(id)+'%',))
                input_image = Image.open((cursor.fetchone())[0])
                inputImageCode = customtkinter.CTkImage(input_image, size=(100, 100))
                label_image = customtkinter.CTkLabel(master=mainpage_right_panel, image=inputImageCode, width=200,
                                                     text="")
                label_image.place(relx=0.6, rely=0.1)
            # customtkinter.CTkLabel(master = mainpage_right_panel,text="Mã hình ảnh").place(relx=0.1,rely = 0.9)
            chenAnh(id)

        # Score's Panel
        def change_to_score():
            for widget in mainpage_right_panel.winfo_children():
                widget.destroy()

            global idMaBangDiem

            detail_label = Label(mainpage_right_panel, text="BẢNG ĐIỂM", font=('Arial', 20, 'bold'), fg='#CFE2F3',
                                 bg="#1f1f1f")
            detail_label.grid(row=0, padx=390)

            table_frame = Frame(mainpage_right_panel)
            table_frame.place(relx=0.5, rely=0.5, anchor=CENTER, width=900, height=550)

            y_scroll = Scrollbar(table_frame, orient=VERTICAL)
            x_scroll = Scrollbar(table_frame, orient=HORIZONTAL)

            style = ttk.Style(mainpage_right_panel)

            style.theme_use("clam")

            # Configure the Treeview style
            style.configure("Treeview", background="#1f1f1f", fieldbackground="#1f1f1f", foreground="white", font=15)
            style.configure("Treeview.Heading", background="#393939", foreground="white", font=15)
            style.map("Treeview.Heading", background=[('active', '#393939',)], foreground=[('active', 'white')])
            style.configure('Vertical.TScrollbar', background="#1f1f1f", troughcolor="#393939", gripcount=0,
                            gripwidth=10,
                            gripperlength=20)
            style.configure('Horizontal.TScrollbar', background="#1f1f1f", troughcolor="#393939", gripcount=0,
                            gripwidth=10,
                            gripperlength=20)

            score_tree = ttk.Treeview(table_frame, yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set,
                                      style="Treeview")

            y_scroll.config(command=score_tree.yview)
            x_scroll.config(command=score_tree.xview)

            y_scroll.pack(side=RIGHT, fill=Y)
            x_scroll.pack(side=BOTTOM, fill=X)

            score_tree['columns'] = (
            "Mã Môn Học", "Tên Môn", "Học Kì", "Điểm Trung Bình", "Điểm Giữa Kì", "Điểm Cuối Kì", "Trạng Thái")

            score_tree.column("Mã Môn Học", anchor=CENTER, width=120, minwidth=120)
            score_tree.column("Tên Môn", anchor=CENTER, width=200, minwidth=200)
            score_tree.column("Học Kì", anchor=CENTER, width=100, minwidth=100)
            score_tree.column("Điểm Trung Bình", anchor=CENTER, width=200, minwidth=200)
            score_tree.column("Điểm Giữa Kì", anchor=CENTER, width=150, minwidth=150)
            score_tree.column("Điểm Cuối Kì", anchor=CENTER, width=150, minwidth=150)
            score_tree.column("Trạng Thái", anchor=CENTER, width=150, minwidth=150)

            score_tree.heading("Mã Môn Học", anchor=CENTER, text="Mã Môn Học")
            score_tree.heading("Tên Môn", anchor=CENTER, text="Tên Môn")
            score_tree.heading("Học Kì", anchor=CENTER, text="Học Kì")
            score_tree.heading("Điểm Trung Bình", anchor=CENTER, text="Điểm Trung Bình")
            score_tree.heading("Điểm Giữa Kì", anchor=CENTER, text="Điểm Giữa Kì")
            score_tree.heading("Điểm Cuối Kì", anchor=CENTER, text="Điểm Cuối Kì")
            score_tree.heading("Trạng Thái", anchor=CENTER, text="Trạng Thái")
            rows = managing_function.get_bangDiem_data(cursor, str(idMaBangDiem))
            for row in rows:
                score_tree.insert('', 'end', values=(row[0], row[1], row[2], ((int(row[4]))+(int(row[5])))/2, row[4], row[5], row[6]))

            score_tree['show'] = 'headings'

            score_tree.pack(fill=BOTH, expand=1)

        # Setting's/Admin's Panel
        def change_to_setting():
            global admin_window
            admin_window = True
            window.quit()
            window.destroy()

        # Main Screen
        def change_to_main_screen():
            for widget in mainpage_right_panel.winfo_children():
                widget.destroy()

            main_label = Label(mainpage_right_panel, text="TRANG CHỦ QUẢN LÝ SINH VIÊN", font=('Arial', 30, 'bold'), fg='#CFE2F3', bg="#1f1f1f", image=mainpanel_logo, compound='top')
            main_label.place(relx=0.5, rely=0.5, anchor=CENTER)

        # Call Data
        global data
        data = read_excel_data()
        connection, cursor = managing_function.connect_to_database()
        # students_data = managing_function.get_students_data(cursor)
        # Others
        # window.after(1000,change_to_main_screen())
        change_to_main_screen()
        window.protocol("WM_DELETE_WINDOW", on_close)
        window.mainloop()

    def LOGIN():
        def on_close():
            result = messagebox.askyesno("Exit", "Bạn muốn thoát ứng dụng?")
            if result:
                window.quit()
                window.destroy()
                sys.exit(0)
        global img1
        global loggedin
        window = customtkinter.CTk()
        window.title('Login')
        window.geometry('800x500+350+120')
        window.configure(bg="black")
        window.resizable(False, False)

        def run_face_detection():
            Recognition_button.configure(state='disabled')
            global verify
            global username
            global id
            global idMaBangDiem
            global idCoVan
            loading_text = StringVar()
            loading_icon = [".",". .",". . ."]
            loading_label = customtkinter.CTkLabel(window, textvariable=loading_text, bg_color='#393939',
                                                   fg_color='#393939',text_color="white", width=100)
            loading_label.place(relx=0.8, rely=0.588, anchor=CENTER)
            for i in range(len(loading_icon)):
                loading_text.set(f'Loading{loading_icon[i]}')
                window.update()
                window.after(500)
            verification = ''
            os.system('face_recognition_face_detecting.py')
            # subprocess.run(["python","face_recognition_face_detecting.py"])
            # os.system('cnn_face_detecting.py')
            loading_label.destroy()
            Recognition_button.configure(state="normal")
            try:
                with open('result.txt', 'r') as f:
                    verification = f.read()
                    id = verification
                connection, cursor = managing_function.connect_to_database()
                cursor.execute("SELECT hovaten FROM sinhvien WHERE idSinhVien = %s", (verification,))
                username_found = cursor.fetchone()
                username = username_found[0]
                cursor.execute("SELECT macv FROM sinhvien WHERE idSinhVien = %s", (verification,))
                idCoVan_found = cursor.fetchone()
                idCoVan = idCoVan_found[0]
                cursor.execute("SELECT mabangdiem FROM sinhvien WHERE idSinhVien = %s", (verification,))
                idMaBangDiem_found = cursor.fetchone()
                idMaBangDiem = idMaBangDiem_found[0]
                with open('result.txt', 'w') as f:
                    f.write('')
            except Exception as e:
                username = ''
                id = None
                idCoVan = None
                idMaBangDiem = None
                print("Từ chối vào ứng dụng!")
                pass
            window.grab_release()
            window.configure(state="normal")
            print(username)
            if username:
                verify = True
                window.after(100, exit_login_tk)

        def on_face_detecting_click():
            window.configure(state="disabled")
            window.grab_set()
            Thread(target=run_face_detection).start()

        def signin():
            global verify
            global username
            global password
            global id
            global idMaBangDiem
            global idCoVan
            username = str(user.get())
            password = str(code.get())
            connection, cursor = managing_function.connect_to_database()
            cursor.execute("SELECT idSinhVien FROM sinhvien WHERE idSinhVien = %s", (username,))
            login_verify = cursor.fetchone()
            if login_verify:
                username_verify = login_verify[0]
                password_verify = login_verify[0]
                if username == str(username_verify) and password == str(password_verify):
                    cursor.execute("SELECT hovaten FROM sinhvien WHERE idSinhVien = %s", (username,))
                    temp = cursor.fetchone()
                    username = temp[0]
                    id = login_verify[0]
                    cursor.execute("SELECT macv FROM sinhvien WHERE idSinhVien = %s", (id,))
                    idCoVan_found = cursor.fetchone()
                    idCoVan = idCoVan_found[0]
                    cursor.execute("SELECT mabangdiem FROM sinhvien WHERE idSinhVien = %s", (id,))
                    idMaBangDiem_found = cursor.fetchone()
                    idMaBangDiem = idMaBangDiem_found[0]
                    # print(id)
                    excel_attandance.diary_attandance(username,id)
                    verify = True
                    exit_login_tk()
            if username == 'admin' and password == 'admin':
                verify = True
                exit_login_tk()
            if verify == False:
                messagebox.showerror("Invalid", "Sai Username hoặc Password")


        img1 = PhotoImage(file='img/login(1).png')
        Label(window, image=img1, bg="#393939").place(x=100, y=150)

        frame = Frame(window, width=350, height=350, bg="#393939", relief=RIDGE , bd=2)
        frame.place(x=550, y=120)

        heading = Label(frame, text='Log In', fg='white', bg='#393939', font=('Microsoft YaHei UI Light', 23, 'bold'))

        heading.place(x=120, y=5)

        ##################################

        def on_enter(e):
            global user
            usernameFrame = Frame(frame, width=295, height=2, bg='white').place(x=25, y=107)
            user = customtkinter.CTkEntry(frame, width=235, height=12, fg_color='#393939', text_color='white',
                                          border_width=0, bg_color='#393939', font=('Microsoft YaHei UI Light', 15),
                                          placeholder_text='Username', placeholder_text_color='white',textvariable=StringVar())

        def on_leave(e):
            global user
            name = user.get()
            if name == '':
                user = customtkinter.CTkEntry(frame, width=235,height=12, fg_color='#393939',text_color='white',
                                              border_width=0, bg_color='#393939', font=('Microsoft YaHei UI Light', 15),placeholder_text='Username',
                                              placeholder_text_color='white')
            usernameFrame = Frame(frame, width=295, height=2, bg='black').place(x=25, y=107)

        user = customtkinter.CTkEntry(frame, width=235,height=12, fg_color='#393939',text_color='white',
                                      border_width=0, bg_color='#393939', font=('Microsoft YaHei UI Light', 15),placeholder_text='Username',
                                      placeholder_text_color='white')
        user.place(x=20, y=60)
        user.bind('<FocusIn>', on_enter)
        user.bind('<FocusOut>', on_leave)

        usernameFrame = Frame(frame, width=295, height=2, bg='black').place(x=25, y=107)

        ##################################
        def on_enter(e):
            global code
            passwordFrame = Frame(frame, width=295, height=2, bg='white').place(x=25, y=177)
            code = customtkinter.CTkEntry(frame, show='*', width=235, height=12, fg_color='#393939', text_color='white',
                                          border_width=0, bg_color='#393939', font=('Microsoft YaHei UI Light', 12),
                                          textvariable=StringVar(),  placeholder_text='Password',
                                          placeholder_text_color='white' )
        def on_leave(e):
            global code
            name = code.get()
            if name == '':
                code = customtkinter.CTkEntry(frame, width=235, height=12, fg_color='#393939', text_color='white',
                                              border_width=0, bg_color='#393939', font=('Microsoft YaHei UI Light', 12),
                                               placeholder_text='Password',
                                              placeholder_text_color='white' )
            passwordFrame = Frame(frame, width=295, height=2, bg='black').place(x=25, y=177)

        code = customtkinter.CTkEntry(frame,show='*', width=235,height=12, fg_color='#393939',
                                      text_color='white', border_width=0, bg_color='#393939', font=('Microsoft YaHei UI Light', 15),
                                      placeholder_text='Password',placeholder_text_color='white')
        code.place(x=20, y=117)
        code.bind('<FocusIn>', on_enter)
        code.bind('<FocusOut>', on_leave)

        passwordFrame = Frame(frame, width=295, height=2, bg='black').place(x=25, y=177)

        #################################

        login_button = customtkinter.CTkButton(frame, width=115, height=35, text='Log in', bg_color='#393939',
                                               fg_color='#393939',text_color='white',border_color='black',hover_color='grey', border_width=2, command=signin)
        login_button.place(x=20,y=180)
        Recognition_button = customtkinter.CTkButton(frame, width=115, height=35, text='Recognition', bg_color='#393939',
                                                     fg_color='#393939', text_color='white',border_color='black',
                                                     hover_color='grey', border_width=2,command=on_face_detecting_click)
        Recognition_button.place(x=140, y=180)
        # label = Label(frame, text="Don't have an account?", fg='black', bg='white', font=('Microsoft YaHei UI Light', 9))
        # label.place(x=75, y=270)
        # sign_up = Button(frame, width=6, text='Sign up', border=0, bg='#393939', cursor='hand2', fg='white')
        # sign_up.place(x=215, y=270)

        def exit_login_tk():
            window.quit()
            window.destroy()
        window.protocol("WM_DELETE_WINDOW", on_close)
        window.mainloop()


    def admin():
        import scanning
        import test_detection
        import customtkinter
        import tkinter
        from tkinter import ttk, messagebox
        from PIL import Image
        import connection
        import matplotlib.pyplot as plt
        import numpy as np
        from matplotlib.figure import Figure
        from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg)
        global rightSideFrame
        def on_close():
            result = messagebox.askyesno("Exit", "Thoát chế độ Admin?")
            if result:
                global value
                value = True
                # print(value)
                # print("Value has changed to True")
                window.quit()
                window.destroy()


        def clear_frame():
            for widgets in rightSideFrame.winfo_children():
                widgets.destroy()

        def callStudentGUI():
            clear_frame()
            rightSideFrame.update()
            StudentFrame = customtkinter.CTkFrame(rightSideFrame, height=300, width=600)
            columns = (1, 2, 3, 4, 5, 6, 7, 8)
            tree = ttk.Treeview(StudentFrame, columns=columns, show="headings", height=17)
            tree.heading(1, text="ID Sinh viên")
            tree.heading(2, text="Họ Và Tên")
            tree.heading(3, text="Email")
            tree.heading(4, text="Mã Lớp")
            tree.heading(5, text="Ngành")
            tree.heading(6, text="Khoa")
            tree.heading(7, text="Hệ Đào Tạo")
            tree.heading(8, text="Niên Khóa")
            mycursor = connection.mydb.cursor()
            sql = """select sv.idSinhVien, sv.hovaten, sv.email,  KhoaHoc.lop, KhoaHoc.nganh, KhoaHoc.khoa, KhoaHoc.hedaotao, KhoaHoc.nienkhoa
                    from SinhVien as sv
                    join KhoAnh 
                    join NguoiCoVan 
                    join KhoaHoc
                    join ThanhTich
                    where sv.mahinhanh = KhoAnh.mahinhanh 
                    and sv.macv = NguoiCoVan.macv 
                    and sv.makhoahoc = KhoaHoc.makhoahoc
                    and sv.mabangdiem = ThanhTich.mabangdiem"""
            mycursor.execute(sql)
            myresult = mycursor.fetchall()
            for x in myresult:
                tree.insert('', "end", iid=x[0], values=x)

            tree.column(1, width=10)
            tree.column(2, width=100)
            tree.column(3, width=150)
            tree.column(4, width=50)
            tree.column(5, width=130)
            tree.column(6, width=130)
            tree.column(7, width=50)
            tree.column(8, width=100)

            def deleteStudent():
                selected = tree.selection()[0]
                query = "delete from SinhVien where idSinhVien = %s"
                data = (selected,)
                delCursor = connection.mydb.cursor()
                yesOrNo = messagebox.askyesno(title="Xóa sinh viên", message="Bạn có chắc muốn xóa sinh viên này không ?")
                if yesOrNo:
                    result = delCursor.execute(query, data)
                    if result != 0:
                        connection.mydb.commit()
                        tree.delete(selected)
                        messagebox.showinfo("Thành công", "Bạn đã xóa sinh viên thành công")
                        callStudentGUI()
                    else:
                        messagebox.showerror("Lỗi", "Có lỗi xảy ra, Vui lòng kiểm tra lại")
                else:
                    messagebox.showerror("Lỗi", "Bạn đã không xóa sinh viên này")

            def editStudent():
                def editStudentFunc():
                    selected = tree.selection()[0]
                    data = (inputName.get(), inputDate.get(), gioi_tinh.get(), inputPhone.get(), inputCccd.get(),
                            inputEmail.get(),
                            inputBorn.get(), inputNation.get(),
                            inputHousehold.get(), int(inputAdsiderCode.get()), int(inputScoreCode.get()),
                            int(inputImageCode.get()),
                            selected)
                    sql = """UPDATE SinhVien SET 
                                hovaten = %s,
                                ngaysinh = %s,
                                gioitinh = %s, 
                                sodienthoai = %s,
                                chungminhthu = %s,
                                email =%s,
                                noisinh =%s,
                                dantoc = %s,
                                hokhau = %s,
                                macv =%s,
                                mabangdiem = %s,
                                mahinhanh =%s
                                WHERE idSinhVien = %s
                                """
                    editCursor = connection.mydb.cursor()
                    result = editCursor.execute(sql, data)
                    if result != 0:
                        messagebox.showinfo("Thành công", "Bạn đã sửa thông tin sinh viên thành công")
                        connection.mydb.commit()
                        callStudentGUI()
                    else:
                        messagebox.showerror("Lỗi", "Có lỗi xảy ra, Vui lòng kiểm tra lại")

                def getCurrentStudent():
                    selected = tree.selection()[0]
                    editCursor = connection.mydb.cursor()
                    sql = "SELECT * FROM SinhVien WHERE idSinhVien = %s"
                    data = (selected,)

                    editCursor.execute(sql, data)
                    result = editCursor.fetchone()
                    return result

                currentStudent = getCurrentStudent()

                editStudentFrame = customtkinter.CTkToplevel(master=StudentFrame, width=500, height=800,
                                                             fg_color="white")
                editStudentFrame.title("Sửa thông tin sinh viên")
                #id
                inputName = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="ID",
                                                   width=100)
                inputName.insert(0, currentStudent[0])
                inputName.place(relx=0.1, rely=0.05)

                #Name
                customtkinter.CTkLabel(master=editStudentFrame, text="Tên sinh viên").place(relx=0.1, rely=0.1)
                inputName = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="Nhập đầy đủ họ và tên...",
                                                   width=200)
                inputName.insert(0, currentStudent[1])
                inputName.place(relx=0.4, rely=0.1)

                # date
                customtkinter.CTkLabel(master=editStudentFrame, text="Ngày tháng năm sinh").place(relx=0.1, rely=0.2)
                inputDate = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="định dạng yyyy-mm-dd",
                                                   width=180)
                inputDate.insert(0, currentStudent[2])
                inputDate.place(relx=0.4, rely=0.2)

                # gender
                gioi_tinh = tkinter.StringVar(value=currentStudent[3])
                nam_radiobutton = customtkinter.CTkRadioButton(
                    editStudentFrame,
                    text="Nam",
                    variable=gioi_tinh,
                    value="nam",
                    width=15,
                    height=15
                )
                nam_radiobutton.place(relx=0.5, rely=0.25)
                nu_radiobutton = customtkinter.CTkRadioButton(
                    editStudentFrame,
                    text="Nữ",
                    variable=gioi_tinh,
                    value="nữ",
                    width=15,
                    height=15
                )
                nu_radiobutton.place(relx=0.7, rely=0.25)
                # cccd
                customtkinter.CTkLabel(master=editStudentFrame, text="Số cccd").place(relx=0.1, rely=0.3)
                inputCccd = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="Nhập số cccd", width=180)
                inputCccd.insert(0, currentStudent[5])
                inputCccd.place(relx=0.4, rely=0.3)

                # phone Number
                customtkinter.CTkLabel(master=editStudentFrame, text="Số điện thoại").place(relx=0.1, rely=0.4)
                inputPhone = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="Nhập số điện thoại",
                                                    width=180)
                inputPhone.insert(0, currentStudent[4])
                inputPhone.place(relx=0.4, rely=0.4)

                # email
                customtkinter.CTkLabel(master=editStudentFrame, text="Email").place(relx=0.1, rely=0.5)
                inputEmail = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="Nhập email", width=180)
                inputEmail.insert(0, currentStudent[6])
                inputEmail.place(relx=0.4, rely=0.5)

                # Place Born
                customtkinter.CTkLabel(master=editStudentFrame, text="Nơi sinh").place(relx=0.1, rely=0.6)
                inputBorn = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="Nhập nơi sinh", width=180)
                inputBorn.insert(0, currentStudent[7])
                inputBorn.place(relx=0.4, rely=0.6)

                # nation
                customtkinter.CTkLabel(master=editStudentFrame, text="Dân tộc").place(relx=0.1, rely=0.65)
                inputNation = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="là dân tộc ...",
                                                     width=180)
                inputNation.insert(0, currentStudent[8])
                inputNation.place(relx=0.4, rely=0.65)

                # Household
                customtkinter.CTkLabel(master=editStudentFrame, text="Hộ khẩu").place(relx=0.1, rely=0.7)
                inputHousehold = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="hộ khẩu ở ...",
                                                        width=200)
                inputHousehold.insert(0, currentStudent[10])
                inputHousehold.place(relx=0.4, rely=0.7)

                # adsider
                customtkinter.CTkLabel(master=editStudentFrame, text="Mã người cố vấn").place(relx=0.1, rely=0.75)
                inputAdsiderCode = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="mã người cố vấn ",
                                                          width=160)
                inputAdsiderCode.insert(0, currentStudent[12])
                inputAdsiderCode.place(relx=0.4, rely=0.75)

                # score
                customtkinter.CTkLabel(master=editStudentFrame, text="Mã bảng điểm").place(relx=0.1, rely=0.8)
                inputScoreCode = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="mã bảng điểm ",
                                                        width=160)
                inputScoreCode.insert(0, currentStudent[14])
                inputScoreCode.place(relx=0.4, rely=0.8)

                # image

                customtkinter.CTkLabel(master=editStudentFrame, text="Mã hình ảnh").place(relx=0.1, rely=0.85)
                inputImageCode = customtkinter.CTkEntry(master=editStudentFrame, placeholder_text="mã hình ảnh ", width=160)
                inputImageCode.insert(0, currentStudent[15])

                inputImageCode.place(relx=0.4, rely=0.85)

                submitButton = customtkinter.CTkButton(master=editStudentFrame, text="OK", command=editStudentFunc).place(
                    relx=0.45,
                    rely=0.95)

            def addStudent():

                def addStudentFunc():
                    data = (
                    int(inputid.get()), inputName.get(), inputDate.get(), gioi_tinh.get(), inputPhone.get(), inputCccd.get(), inputEmail.get(),
                    inputBorn.get(), inputNation.get(),
                    inputHousehold.get(), int(inputAdsiderCode.get()), int(inputClassCode.get()) ,int(inputScoreCode.get()), int(inputImageCode.get()))
                    sql = """
                            insert into SinhVien(idsinhvien, hovaten, ngaysinh, gioitinh, sodienthoai, chungminhthu, email,noisinh, dantoc, tongiao, hokhau, trangthai, macv, makhoahoc, mabangdiem, mahinhanh)
                            value (%s, %s,date(%s), %s, %s, %s, %s, %s, %s, 'khong', %s, 1, %s, %s, %s, %s)
                        """
                    result = mycursor.execute(sql, data)
                    connection.mydb.commit()
                    if result != 0:
                        messagebox.showinfo("Thành công", "Bạn đã thêm sinh viên thành công")
                        callStudentGUI()
                    else:
                        messagebox.showerror("Lỗi", "có lỗi xảy ra, Vui lòng xem xét lại các trường đã nhập vào")

                addStudentFrame = customtkinter.CTkToplevel(master=StudentFrame, width=500, height=800, fg_color="white")
                addStudentFrame.title("Thêm sinh viên")
                customtkinter.CTkLabel(master=addStudentFrame, text="Chú ý điền cho chuẩn", text_color="red",
                                       font=('Times New Roman', 20, 'bold')).place(relx=0.2, rely=0.05)
                # id
                inputid = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="id",width=100)
                inputid.place(relx=0.7, rely=0.05)

                # name
                customtkinter.CTkLabel(master=addStudentFrame, text="Tên sinh viên").place(relx=0.1, rely=0.1)
                inputName = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="Nhập đầy đủ họ và tên...",
                                                   width=200)
                inputName.place(relx=0.4, rely=0.1)

                # date
                customtkinter.CTkLabel(master=addStudentFrame, text="Ngày tháng năm sinh").place(relx=0.1, rely=0.2)
                inputDate = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="định dạng yyyy-mm-dd",
                                                   width=180)
                inputDate.place(relx=0.4, rely=0.2)

                # gender
                gioi_tinh = tkinter.StringVar()
                nam_radiobutton = customtkinter.CTkRadioButton(
                    addStudentFrame,
                    text="Nam",
                    variable=gioi_tinh,
                    value="Nam",
                    width=15,
                    height=15
                )
                nam_radiobutton.place(relx=0.5, rely=0.25)
                nu_radiobutton = customtkinter.CTkRadioButton(
                    addStudentFrame,
                    text="Nữ",
                    variable=gioi_tinh,
                    value="Nữ",
                    width=15,
                    height=15
                )
                nu_radiobutton.place(relx=0.7, rely=0.25)
                # cccd
                customtkinter.CTkLabel(master=addStudentFrame, text="Số cccd").place(relx=0.1, rely=0.3)
                inputCccd = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="Nhập số cccd", width=180)
                inputCccd.place(relx=0.4, rely=0.3)

                # phone Number
                customtkinter.CTkLabel(master=addStudentFrame, text="Số điện thoại").place(relx=0.1, rely=0.4)
                inputPhone = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="Nhập số điện thoại",
                                                    width=180)
                inputPhone.place(relx=0.4, rely=0.4)

                # email
                customtkinter.CTkLabel(master=addStudentFrame, text="Email").place(relx=0.1, rely=0.5)
                inputEmail = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="Nhập email", width=180)
                inputEmail.place(relx=0.4, rely=0.5)

                # Place Born
                customtkinter.CTkLabel(master=addStudentFrame, text="Nơi sinh").place(relx=0.1, rely=0.6)
                inputBorn = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="Nhập nơi sinh", width=180)
                inputBorn.place(relx=0.4, rely=0.6)

                # nation
                customtkinter.CTkLabel(master=addStudentFrame, text="Dân tộc").place(relx=0.1, rely=0.65)
                inputNation = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="là dân tộc ...", width=180)
                inputNation.place(relx=0.4, rely=0.65)

                # Household
                customtkinter.CTkLabel(master=addStudentFrame, text="Hộ khẩu").place(relx=0.1, rely=0.7)
                inputHousehold = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="hộ khẩu ở ...", width=200)
                inputHousehold.place(relx=0.4, rely=0.7)

                # adsider
                customtkinter.CTkLabel(master=addStudentFrame, text="Mã người cố vấn").place(relx=0.1, rely=0.75)
                inputAdsiderCode = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="mã người cố vấn ",
                                                          width=160)
                inputAdsiderCode.place(relx=0.4, rely=0.75)

                # score
                customtkinter.CTkLabel(master=addStudentFrame, text="Mã bảng điểm").place(relx=0.1, rely=0.8)
                inputScoreCode = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="mã bảng điểm ", width=160)
                inputScoreCode.place(relx=0.4, rely=0.8)

                # image
                customtkinter.CTkLabel(master=addStudentFrame, text="Mã hình ảnh").place(relx=0.1, rely=0.85)
                inputImageCode = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="mã hình ảnh ", width=160)
                inputImageCode.place(relx=0.4, rely=0.85)

                # Class
                customtkinter.CTkLabel(master=addStudentFrame, text="Mã khóa học").place(relx=0.1, rely=0.9)
                inputClassCode = customtkinter.CTkEntry(master=addStudentFrame, placeholder_text="mã khóa học ",width=160)
                inputClassCode.place(relx=0.4, rely=0.9)

                submitButton = customtkinter.CTkButton(master=addStudentFrame, text="Add", command=addStudentFunc).place(
                    relx=0.45, rely=0.95)


            addBtn = customtkinter.CTkButton(master=StudentFrame,
                                             width=80,
                                             height=32,
                                             border_width=0,
                                             corner_radius=8,
                                             text="Thêm",
                                             bg_color="white",
                                             command=addStudent)
            addBtn.place(relx=0.3, rely=0.9, anchor=tkinter.CENTER)
            editBtn = customtkinter.CTkButton(master=StudentFrame,
                                              width=80,
                                              height=32,
                                              border_width=0,
                                              corner_radius=8,
                                              text="Sửa",
                                              bg_color="white",
                                              command=editStudent)
            editBtn.place(relx=0.5, rely=0.9, anchor=tkinter.CENTER)
            deleteBtn = customtkinter.CTkButton(master=StudentFrame,
                                                width=80,
                                                height=32,
                                                border_width=0,
                                                corner_radius=8,
                                                text="Xóa",
                                                bg_color="white",
                                                command=deleteStudent)
            deleteBtn.place(relx=0.7, rely=0.9, anchor=tkinter.CENTER)
            tree.pack(padx=10, pady=10)

            StudentFrame.pack(fill=tkinter.BOTH, padx=10, pady=10)

        def callCourseGUI():
            clear_frame()
            CourseFrame = customtkinter.CTkFrame(rightSideFrame, height=200, width=600, fg_color="white")
            tree = ttk.Treeview(CourseFrame, columns=("mamon", "tenmon"), show="headings", height=17)
            tree.heading("mamon", text="Mã Môn Học")
            tree.heading("tenmon", text="Tên Môn Học")
            mycursor = connection.mydb.cursor()
            sql = "SELECT * FROM MonHoc"

            mycursor.execute(sql)

            myresult = mycursor.fetchall()
            for x in myresult:
                tree.insert('', "end", iid=x[0], values=x)
            tree.column("mamon", width=330)
            tree.column("tenmon", width=340)
            tree.pack(padx=10, pady=10)

            def addSubject():
                def addSubjectFunc():
                    data = (inputName.get(),)
                    sql = """
                            INSERT INTO MonHoc (tenmon)
                            value (%s)
                            """
                    result = mycursor.execute(sql, data)
                    if result != 0:
                        messagebox.showinfo("Thành công", "Bạn đã thêm môn học thành công")
                        connection.mydb.commit()
                        callCourseGUI()
                    else:
                        messagebox.showerror("Lỗi", "có lỗi xảy ra, Vui lòng xem xét lại các trường đã nhập vào")

                addSubjectFrame = customtkinter.CTkToplevel(master=CourseFrame, width=400, height=500, fg_color="white")
                customtkinter.CTkLabel(master=addSubjectFrame, text="Tên Môn").place(relx=0.1, rely=0.4)
                inputName = customtkinter.CTkEntry(master=addSubjectFrame, placeholder_text="Nhập tên môn", width=180)
                inputName.place(relx=0.4, rely=0.4)
                submitButton = customtkinter.CTkButton(master=addSubjectFrame, text="Add", command=addSubjectFunc).place(
                    relx=0.45, rely=0.8)

            def editSubject():
                editSubjectFrame = customtkinter.CTkToplevel(master=CourseFrame, width=600, height=300, fg_color="white")
                editSubjectFrame.title("Sửa thông tin môn học")
                customtkinter.CTkLabel(master=editSubjectFrame, text="Tên môn").place(relx=0.2, rely=0.3)
                inputNameSub = customtkinter.CTkEntry(master=editSubjectFrame, placeholder_text="Nhập tên môn ", width=180)
                inputNameSub.place(relx=0.4, rely=0.3)

                def editsubjectFunc():
                    selected = tree.selection()[0]
                    data = (inputNameSub.get(), int(selected))
                    sql = """
                        UPDATE MonHoc 
                        set tenmon = %s 
                        where mamonhoc = %s
                        """
                    if selected:
                        mycursor = connection.mydb.cursor()
                        result = mycursor.execute(sql, data)
                        if result != 0:
                            messagebox.showinfo("Đổi Thông Tin Môn Học", "Đã đổi tên môn thành công")
                            connection.mydb.commit()
                            callCourseGUI()
                        else:
                            messagebox.showerror("Lỗi", "Có lỗi xảy ra với thao tác này")
                    else:
                        messagebox.showinfo("Có vấn đề", "Chưa chọn môn học cần thay đổi ")

                submitBtn = customtkinter.CTkButton(master=editSubjectFrame, text="OK", command=editsubjectFunc).place(
                    relx=0.4, rely=0.8)

            def deleteSubject():
                selected = tree.selection()[0]
                query = "delete from MonHoc where mamonhoc = %s"
                data = (selected,)
                delCursor = connection.mydb.cursor()
                yesOrNo = messagebox.askyesno(title="Xóa môn học", message="Bạn có chắc muốn xóa môn học này không ?")
                if yesOrNo:
                    result = delCursor.execute(query, data)
                    if result != 0:
                        connection.mydb.commit()
                        tree.delete(selected)
                        messagebox.showinfo("Thành công", "Bạn đã xóa môn học thành công")
                        callCourseGUI()
                    else:
                        messagebox.showerror("Lỗi", "Có lỗi xảy ra, Vui lòng kiểm tra lại")
                else:
                    messagebox.showerror("Lỗi", "Bạn đã không xóa môn học này")

            editBtn = customtkinter.CTkButton(master=CourseFrame,
                                              width=80,
                                              height=32,
                                              border_width=0,
                                              corner_radius=8,
                                              text="Thêm",
                                              command=addSubject)
            editBtn.place(relx=0.3, rely=0.9, anchor=tkinter.CENTER)
            editBtn = customtkinter.CTkButton(master=CourseFrame,
                                              width=80,
                                              height=32,
                                              border_width=0,
                                              corner_radius=8,
                                              text="Sửa",
                                              command=editSubject)
            editBtn.place(relx=0.5, rely=0.9, anchor=tkinter.CENTER)
            deleteBtn = customtkinter.CTkButton(master=CourseFrame,
                                                width=80,
                                                height=32,
                                                border_width=0,
                                                corner_radius=8,
                                                text="Xóa",
                                                command=deleteSubject)
            deleteBtn.place(relx=0.7, rely=0.9, anchor=tkinter.CENTER)

            CourseFrame.pack(fill=tkinter.BOTH, padx=10, pady=10)

        def callDashboardGUI():
            clear_frame()
            DashboardFrame = customtkinter.CTkFrame(rightSideFrame, height=500, width=600, fg_color="white")
            graphLeftFrame = customtkinter.CTkFrame(master=DashboardFrame, width=200, height=500, fg_color="black")
            graphRightFrame = customtkinter.CTkFrame(master=DashboardFrame, width=280, height=500, fg_color="blue")
            # Biểu đồ bên trái
            # Tạo dữ liệu

            mycursor = connection.mydb.cursor()
            sql_diemcuoiki = """select count(mabangdiem)
                                from bangdiemchitiet as ct
                                where ct.diemcuoiki = '%s';"""

            query_diem = """SELECT count(diem)
                                    from thanhtich
                                    where thanhtich.diem >= '%s'
                                    and thanhtich.diem < '%s';"""
            # 8.5

            thangdiem = [6, 7.9, 8.5]
            mycursor.execute(query_diem, (thangdiem[0], thangdiem[1]))
            tong_diem_trung_binh = mycursor.fetchone()[0]
            mycursor.execute(query_diem, (thangdiem[1], thangdiem[2]))
            tong_diem_kha = mycursor.fetchone()[0]
            mycursor.execute(query_diem, (thangdiem[2], 10))
            tong_diem_gioi = mycursor.fetchone()[0]

            so_diemcuoiki = []
            for i in range(7, 11):
                data = (i,)
                mycursor.execute(sql_diemcuoiki, data)
                result = mycursor.fetchone()[0]
                so_diemcuoiki.append(result)

            mocdiem = np.array(["7", "8", "9", "10"])
            trucdiem = np.array(so_diemcuoiki)
            fig, ax = plt.subplots(figsize=(4, 3))
            ax.bar(mocdiem, trucdiem, color="#4CAF50")
            ax.set_ylabel('Thang điểm')
            ax.set_title('Số điểm cuối kì')
            # Thêm biểu đồ vào frame
            canvas = FigureCanvasTkAgg(fig, master=graphLeftFrame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tkinter.Y, expand=False)

            # Biển đồ bên phải
            # Dữ liệu cho biểu đồ pie
            y_pie = np.array([tong_diem_trung_binh, tong_diem_kha, tong_diem_gioi])
            mylabels = ["TB", "Khá", "Giỏi"]
            # Kích thước của biểu đồ
            fig_width = 6
            fig_height = 4
            # Vẽ biểu đồ pie
            fig, ax = plt.subplots(figsize=(fig_width, fig_height))
            ax.pie(y_pie, labels=mylabels)
            ax.set_title('Điểm trung bình')

            # Thêm biểu đồ vào frame
            canvas = FigureCanvasTkAgg(fig, master=graphRightFrame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tkinter.BOTH, expand=True)

            graphLeftFrame.pack(side=tkinter.LEFT, padx=10, pady=10)
            graphRightFrame.pack(side=tkinter.RIGHT, padx=10, pady=10)
            DashboardFrame.pack(fill=tkinter.BOTH, padx=10, pady=10)
            window.update()

        def callUpdateClass():
            # connect db
            db = mysql.connector.connect(user='root', password='', host='localhost', database='qlthongtinsv')
            top = customtkinter.CTk()
            cursor = db.cursor()
            # top.title('GUI')
            # top.geometry("1000x1000")
            # top.attributes('-topmost', True)
            query = 'select * from khoahoc'
            cursor.execute(query)
            rows = cursor.fetchall()
            # top.resizable(False, False) # khong dc phong to cua so

            ############
            info = customtkinter.CTkToplevel(master=window, width=1110, height=600, fg_color="white")
            info.title("Class")
            # listbox = tkinter.Listbox(info, width=150, height=40).grid(row=1, columnspan=2)

            #
            customtkinter.CTkLabel(master=info, text="Quản Lý Khoá Học ", text_color="#E3CF57",
                                   font=('Times New Roman', 20, 'bold')).place(relx=0.3, rely=0.05)
            # customtkinter.CTkLabel(master = info, text="Thông tin môn học", text_color="red", font=('Times New Roman', 20, 'bold')).place(relx = 0.1, rely = 0.1)
            # id giang vien
            customtkinter.CTkLabel(master=info, text="Mã khóa học: ").place(relx=0.05, rely=0.2)
            entry1 = customtkinter.CTkEntry(master=info, placeholder_text=None)
            entry1.place(relx=0.15, rely=0.2)

            # ten giang vien
            customtkinter.CTkLabel(master=info, text="Mã ngành học: ").place(relx=0.05, rely=0.25)
            entry2 = customtkinter.CTkEntry(master=info, placeholder_text=None)
            entry2.place(relx=0.15, rely=0.25)
            # Email
            customtkinter.CTkLabel(master=info, text="Lớp: ").place(relx=0.05, rely=0.3)
            entry3 = customtkinter.CTkEntry(master=info, placeholder_text=None)
            entry3.place(relx=0.15, rely=0.3)
            # So dien thoai
            customtkinter.CTkLabel(master=info, text="Ngành: ").place(relx=0.05, rely=0.35)
            entry4 = customtkinter.CTkEntry(master=info, placeholder_text=None)
            entry4.place(relx=0.15, rely=0.35)
            ##############
            customtkinter.CTkLabel(master=info, text="Khoa: ").place(relx=0.3, rely=0.2)
            entry5 = customtkinter.CTkEntry(master=info, placeholder_text=None)
            entry5.place(relx=0.40, rely=0.2)

            customtkinter.CTkLabel(master=info, text="Hệ đào tạo: ").place(relx=0.3, rely=0.25)
            entry6 = customtkinter.CTkEntry(master=info, placeholder_text=None)
            entry6.place(relx=0.40, rely=0.25)

            customtkinter.CTkLabel(master=info, text="Niên khóa: ").place(relx=0.3, rely=0.3)
            entry7 = customtkinter.CTkEntry(master=info, placeholder_text=None)
            entry7.place(relx=0.40, rely=0.3)

            # Tree view
            tree = ttk.Treeview(info, columns=(1, 2, 3, 4, 5, 6, 7), show='headings', height='5')
            tree.column(1, width=145)
            tree.column(2, width=145)
            tree.column(3, width=145)
            tree.column(4, width=145)
            tree.column(5, width=145)
            tree.column(6, width=145)
            tree.column(7, width=145)
            tree.heading(1, text='Mã khóa học')
            tree.heading(2, text='Mã ngành học')
            tree.heading(3, text='Lớp')
            tree.heading(4, text='Ngành')
            tree.heading(5, text='Khoa')
            tree.heading(6, text='Hệ đào tạo')
            tree.heading(7, text='Niên khóa')
            tree.place(x=50, y=340)

            ################# Load database
            for i in rows:
                tree.insert('', 'end', iid=i[0], values=i)

            ################# Delete

            ###
            def delete():
                selected = tree.selection()[0]
                query = 'DELETE FROM khoahoc WHERE makhoahoc = %s'
                data = (selected,)
                q_del = db.cursor()
                q_del.execute(query, data)
                db.commit()
                tree.delete(selected)
                window.update()

            customtkinter.CTkLabel(master=info, text="Tìm kiếm theo ").place(relx=0.3, rely=0.45)
            # Combobox ID
            idOption = ["Mã Ngành", "Lớp", "Ngành/Khoa"]
            my_combo = customtkinter.CTkComboBox(info, values=idOption)
            my_combo.place(relx=0.4, rely=0.45)
            class_entry = customtkinter.CTkEntry(info, placeholder_text= "", placeholder_text_color='gray', width=140)
            class_entry.place(relx=0.6, rely=0.45)
            if my_combo.get() == "Mã Ngành":
                class_entry.configure(placeholder_text="Nhập Mã Ngành")
            if my_combo.get() == "Lớp":
                class_entry.configure(placeholder_text="Nhập Lớp")
            if my_combo.get() == "Ngành/Khoa":
                class_entry.configure(placeholder_text="Nhập Ngành/Khoa")



            ####################

            #########
            def select_record(e):
                # clear entry boxes
                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')
                entry5.delete(0, 'end')
                entry6.delete(0, 'end')
                entry7.delete(0, 'end')
                # grad record Number
                selected = tree.focus()
                values = tree.item(selected, 'values')
                entry1.insert(0, values[0])
                entry2.insert(0, values[1])
                entry3.insert(0, values[2])
                entry4.insert(0, values[3])
                entry5.insert(0, values[4])
                entry6.insert(0, values[5])
                entry7.insert(0, values[6])

            tree.bind("<ButtonRelease - 1>", select_record)

            ####################
            def clear_entries():
                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')
                entry5.delete(0, 'end')
                entry6.delete(0, 'end')
                entry7.delete(0, 'end')

            ###################
            def update_record():
                selected = tree.focus()
                tree.item(selected, text="", values=(
                entry1.get(), entry2.get(), entry3.get(), entry4.get(), entry5.get(), entry6.get(), entry7.get()))

                cursor.execute("""
                    UPDATE khoahoc 
                    SET manganhhoc = %s, lop = %s, nganh = %s, khoa = %s, hedaotao = %s, nienkhoa = %s
                    WHERE makhoahoc = %s
                """, (entry2.get(), entry3.get(), entry4.get(), entry5.get(), entry6.get(), entry7.get(), entry1.get()))

                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')
                entry5.delete(0, 'end')
                entry6.delete(0, 'end')
                entry7.delete(0, 'end')

                db.commit()
                window.update()


            ###################
            def add_record1():
                cursor.execute("INSERT INTO khoahoc VALUES (%s, %s, %s, %s,%s,%s,%s)",
                               {
                                   'makhoahoc': entry1.get(),
                                   'manganhhoc': entry2.get(),
                                   'lop': entry3.get(),
                                   'nganh': entry4.get(),
                                   'khoa': entry5.get(),
                                   'hedaotao': entry6.get(),
                                   'nienkhoa': entry7.get(),
                               })
                db.commit()
                window.update()

                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')
                entry5.delete(0, 'end')
                entry6.delete(0, 'end')
                entry7.delete(0, 'end')

            def add_record():
                cursor.execute("INSERT INTO khoahoc VALUES (%s, %s, %s, %s, %s, %s, %s)",
                               (
                                   entry1.get(),
                                   entry2.get(),
                                   entry3.get(),
                                   entry4.get(),
                                   entry5.get(),
                                   entry6.get(),
                                   entry7.get(),
                               ))
                db.commit()
                window.update()

                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')
                entry5.delete(0, 'end')
                entry6.delete(0, 'end')
                entry7.delete(0, 'end')
                for i in rows:
                    if not tree.exists(i[0]):
                        tree.insert('', 'end', iid=i[0], values=i)

            # Mở kết nối đến cơ sở dữ liệu ở đây

            def search_records2():
                try:
                    # Lấy giá trị cần tìm kiếm từ entry
                    lookup_record = class_entry.get()

                    # Xóa dữ liệu hiện có trong treeview
                    for record in tree.get_children():
                        tree.delete(record)

                    # Thực hiện truy vấn SQL để tìm kiếm các bản ghi phù hợp
                    if my_combo.get() == "Mã Ngành":
                        cursor.execute("SELECT * FROM khoahoc WHERE manganhhoc LIKE %s", ('%' + lookup_record + '%',))
                    if my_combo.get() == "Lớp":
                        cursor.execute("SELECT * FROM khoahoc WHERE lop LIKE %s", ('%' + lookup_record + '%',))
                    if my_combo.get() == "Ngành/Khoa":
                        cursor.execute("SELECT * FROM khoahoc WHERE khoa LIKE %s", ('%' + lookup_record + '%',))
                    # cursor.execute("SELECT * FROM khoahoc WHERE makhoahoc LIKE %s", ('%' + lookup_record + '%',))
                    rows = cursor.fetchall()  # Lấy kết quả từ truy vấn SQL
                    # Cập nhật treeview với các bản ghi được tìm thấy
                    for row in rows:
                        tree.insert('', 'end', iid=row[0], values=row)

                    db.commit()
                    window.update()
                except mysql.connector.Error as err:
                    print("Lỗi khi thực hiện truy vấn SQL:", err)

            # Xem tất cả các dòng
            def show_all_records():
                try:
                    # Xóa dữ liệu hiện có trong treeview
                    for record in tree.get_children():
                        tree.delete(record)

                    # Thực hiện truy vấn SQL để lấy tất cả các bản ghi
                    cursor.execute("SELECT * FROM khoahoc")
                    rows = cursor.fetchall()  # Lấy kết quả từ truy vấn SQL

                    # Cập nhật treeview với tất cả các bản ghi
                    for row in rows:
                        tree.insert('', 'end', iid=row[0], values=row)

                    db.commit()
                    window.update()
                except mysql.connector.Error as err:
                    print("Lỗi khi thực hiện truy vấn SQL:", err)

            # remove one records
            def remove_one_record():
                x = tree.selection()[0]
                tree.delete(x)
                cursor.execute("DELETE from khoahoc WHERE makhoahoc = " + entry1.get())
                db.commit()
                window.update()


            ##############################################
            # Button
            button_tim = customtkinter.CTkButton(master=info, width=50, height=25, border_width=0, corner_radius=8,
                                                 text="Tìm Kiếm", command=search_records2)
            button_tim.place(relx=0.85, rely=0.48, anchor=tkinter.CENTER)

            button_xem = customtkinter.CTkButton(master=info, width=60, height=25, border_width=0, corner_radius=8,
                                                 text="Xem tất cả", command=show_all_records)
            button_xem.place(relx=0.95, rely=0.48, anchor=tkinter.CENTER)

            button_them = customtkinter.CTkButton(master=info, width=70, height=25, border_width=0, corner_radius=8,
                                                  text="Thêm", command=add_record)
            button_them.place(relx=0.1, rely=0.45, anchor=tkinter.CENTER)

            button_xoa = customtkinter.CTkButton(master=info, width=68, height=25, border_width=0, corner_radius=8,
                                                 text="Xóa", command=remove_one_record)
            button_xoa.place(relx=0.2, rely=0.45, anchor=tkinter.CENTER)

            button_capnhat = customtkinter.CTkButton(master=info, width=60, height=25, border_width=0, corner_radius=8,
                                                     text="Cập nhật", command=update_record)
            button_capnhat.place(relx=0.1, rely=0.5, anchor=tkinter.CENTER)

            button_lammoi = customtkinter.CTkButton(master=info, width=60, height=25, border_width=0, corner_radius=8,
                                                    text="Làm mới", command=clear_entries)
            button_lammoi.place(relx=0.2, rely=0.5, anchor=tkinter.CENTER)


        def callUpdateCoVan():
            # connect db
            db = mysql.connector.connect(user='root', password='', host='localhost', database='qlthongtinsv')
            cursor = db.cursor()
            query = 'select * from nguoicovan'
            cursor.execute(query)
            rows = cursor.fetchall()
            # top.resizable(False, False) # khong dc phong to cua so
            # idSinhVien = 3

            def getCurrentStudent(idSinhVien):
                editCursor = db.cursor()
                sql = "SELECT * FROM SinhVien WHERE idSinhVien = %s"
                data = (idSinhVien,)

                editCursor.execute(sql, data)
                result = editCursor.fetchone()
                return result

            curentstudent = getCurrentStudent(id)
            print(curentstudent)

            ####

            # idCoVan = 2

            def getCurrentStudent(idCoVan):
                editCursor = db.cursor()
                sql = "SELECT * FROM nguoicovan WHERE macv = %s"
                data = (idCoVan,)

                editCursor.execute(sql, data)
                result = editCursor.fetchone()
                return result

            curentcovan = getCurrentStudent(idCoVan)
            print(curentcovan)
            ####

            idMaBangDiem = 1

            def getCurrentStudent(idMaBangDiem):
                editCursor = db.cursor()
                sql = "SELECT * FROM bangdiemchitiet WHERE mabangdiem = %s"
                data = (idMaBangDiem,)

                editCursor.execute(sql, data)
                result = editCursor.fetchone()
                return result

            curentBangDiem = getCurrentStudent(idMaBangDiem)
            print(curentBangDiem)
            ############
            info = customtkinter.CTkToplevel(master=window, width=1000, height=600, fg_color="white")
            info.title("Adviser")
            # listbox = tkinter.Listbox(info, width=150, height=40).grid(row=1, columnspan=2)

            #
            customtkinter.CTkLabel(master=info, text="Quản Lý Thông Tin Cố Vấn ", text_color="#E3CF57",
                                   font=('Times New Roman', 20, 'bold')).place(relx=0.3, rely=0.05)
            # customtkinter.CTkLabel(master = info, text="Thông tin môn học", text_color="red", font=('Times New Roman', 20, 'bold')).place(relx = 0.1, rely = 0.1)
            # id giang vien
            customtkinter.CTkLabel(master=info, text="Mã cố vấn: ").place(relx=0.05, rely=0.2)
            entry1 = customtkinter.CTkEntry(master=info, placeholder_text=None,width=100)
            entry1.place(relx=0.15, rely=0.2)

            # ten giang vien
            customtkinter.CTkLabel(master=info, text="Tên cố vấn: ").place(relx=0.05, rely=0.25)
            entry2 = customtkinter.CTkEntry(master=info, placeholder_text=None,width=100)
            entry2.place(relx=0.15, rely=0.25)
            # Email
            customtkinter.CTkLabel(master=info, text="Email: ").place(relx=0.05, rely=0.3)
            entry3 = customtkinter.CTkEntry(master=info, placeholder_text=None,width=100)
            entry3.place(relx=0.15, rely=0.3)
            # So dien thoai
            customtkinter.CTkLabel(master=info, text="Số điện thoại: ").place(relx=0.05, rely=0.35)
            entry4 = customtkinter.CTkEntry(master=info, placeholder_text=None,width=100)
            entry4.place(relx=0.15, rely=0.35)

            # Tree view
            tree = ttk.Treeview(info, columns=(1, 2, 3, 4), show='headings', height='5')
            tree.column(1, width=155)
            tree.column(2, width=155)
            tree.column(3, width=155)
            tree.column(4, width=155)
            tree.heading(1, text='Mã cố vấn')
            tree.heading(2, text='Tên cố vấn')
            tree.heading(3, text='Email')
            tree.heading(4, text='Số điện thoại')
            tree.place(x=300, y=115)

            ################# Load database
            for i in rows:
                tree.insert('', 'end', iid=i[0], values=i)

                ################# Delete

            ###
            def delete():
                selected = tree.selection()[0]
                query = 'DELETE FROM nguoicovan WHERE macv = %s'
                data = (selected,)
                q_del = db.cursor()
                q_del.execute(query, data)
                db.commit()
                tree.delete(selected)
                window.update()

            customtkinter.CTkLabel(master=info, text="Tìm kiếm theo tên: ").place(relx=0.3, rely=0.11)
            # Combobox ID
            my_combo = customtkinter.CTkEntry(info,placeholder_text="Nhập tên", placeholder_text_color='gray' ,width=100)
            my_combo.place(relx=0.45, rely=0.11)

            ####################

            #########
            def select_record(e):
                # clear entry boxes
                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')
                # grad record Number
                selected = tree.focus()
                values = tree.item(selected, 'values')
                entry1.insert(0, values[0])
                entry2.insert(0, values[1])
                entry3.insert(0, values[2])
                entry4.insert(0, values[3])

            tree.bind("<ButtonRelease - 1>", select_record)

            ####################
            def clear_entries():
                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')

            ###################
            def update_record():
                selected = tree.focus()
                tree.item(selected, text="", values=(entry1.get(), entry2.get(), entry3.get(), entry4.get()))

                cursor.execute("""
                    UPDATE nguoicovan 
                    SET hovaten = %s, email = %s, sodienthoai = %s
                    WHERE macv = %s
                """, (entry2.get(), entry3.get(), entry4.get(), entry1.get()))

                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')

                db.commit()
                window.update()
                window

            ###################
            def add_record1():
                cursor.execute("INSERT INTO nguoicovan VALUES (%s, %s, %s, %s)",
                               {
                                   'macv': entry1.get(),
                                   'hovaten': entry2.get(),
                                   'email': entry3.get(),
                                   'sodienthoai': entry4.get(),
                               })
                db.commit()
                window
                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')

            def add_record():
                cursor.execute("INSERT INTO nguoicovan VALUES (%s, %s, %s, %s)",
                               (
                                   entry1.get(),
                                   entry2.get(),
                                   entry3.get(),
                                   entry4.get(),
                               ))
                db.commit()
                window.update()
                window
                entry1.delete(0, 'end')
                entry2.delete(0, 'end')
                entry3.delete(0, 'end')
                entry4.delete(0, 'end')
                for i in rows:
                    if not tree.exists(i[0]):
                        tree.insert('', 'end', iid=i[0], values=i)
                    # Mở kết nối đến cơ sở dữ liệu ở đây

            def search_records2():
                try:
                    # Lấy giá trị cần tìm kiếm từ combobox
                    lookup_record = my_combo.get()

                    # Xóa dữ liệu hiện có trong treeview
                    for record in tree.get_children():
                        tree.delete(record)

                    # Thực hiện truy vấn SQL để tìm kiếm các bản ghi phù hợp
                    cursor.execute("SELECT * FROM nguoicovan WHERE hovaten LIKE %s", ('%' + lookup_record + '%',))
                    rows = cursor.fetchall()  # Lấy kết quả từ truy vấn SQL

                    # Cập nhật treeview với các bản ghi được tìm thấy
                    for row in rows:
                        tree.insert('', 'end', iid=row[0], values=row)

                    db.commit()
                    window.update()
                except mysql.connector.Error as err:
                    print("Lỗi khi thực hiện truy vấn SQL:", err)

            # Đóng kết nối đến cơ sở dữ liệu ở đây

            def search_records1():
                # Lấy giá trị cần tìm kiếm từ combobox
                lookup_record = my_combo.get()

                # Xóa dữ liệu hiện có trong treeview
                for record in tree.get_children():
                    tree.delete(record)

                # Thực hiện truy vấn SQL để tìm kiếm các bản ghi phù hợp
                cursor.execute("SELECT * FROM nguoicovan WHERE macv LIKE %s", ('%' + lookup_record + '%',))
                rows = cursor.fetchall()  # Lấy kết quả từ truy vấn SQL

                # Cập nhật treeview với các bản ghi được tìm thấy
                for row in rows:
                    tree.insert('', 'end', iid=row[0], values=row)

                db.commit()
                window.update
                window

            def show_all_records():
                try:
                    # Xóa dữ liệu hiện có trong treeview
                    for record in tree.get_children():
                        tree.delete(record)

                    # Thực hiện truy vấn SQL để lấy tất cả các bản ghi
                    cursor.execute("SELECT * FROM nguoicovan")
                    rows = cursor.fetchall()  # Lấy kết quả từ truy vấn SQL

                    # Cập nhật treeview với tất cả các bản ghi
                    for row in rows:
                        tree.insert('', 'end', iid=row[0], values=row)

                    db.commit()
                    window.update()
                except mysql.connector.Error as err:
                    print("Lỗi khi thực hiện truy vấn SQL:", err)

            ##############################################
            # Button
            button_tim = customtkinter.CTkButton(master=info, width=50, height=25, border_width=0, corner_radius=8,
                                                 text="Tìm Kiếm", command=search_records2)
            button_tim.place(relx=0.65, rely=0.132, anchor=tkinter.CENTER)

            button_xem = customtkinter.CTkButton(master=info, width=60, height=25, border_width=0, corner_radius=8,
                                                 text="Xem tất cả", command=show_all_records)
            button_xem.place(relx=0.75, rely=0.132, anchor=tkinter.CENTER)

            button_them = customtkinter.CTkButton(master=info, width=70, height=25, border_width=0, corner_radius=8,
                                                  text="Thêm", command=add_record)
            button_them.place(relx=0.1, rely=0.45, anchor=tkinter.CENTER)

            button_capnhat = customtkinter.CTkButton(master=info, width=60, height=25, border_width=0, corner_radius=8,
                                                     text="Cập nhật", command=update_record)
            button_capnhat.place(relx=0.1, rely=0.5, anchor=tkinter.CENTER)

            button_xoa = customtkinter.CTkButton(master=info, width=68, height=25, border_width=0, corner_radius=8,
                                                 text="Xóa", command=delete)
            button_xoa.place(relx=0.2, rely=0.45, anchor=tkinter.CENTER)

            button_lammoi = customtkinter.CTkButton(master=info, width=60, height=25, border_width=0, corner_radius=8,
                                                    text="Làm mới", command=clear_entries)
            button_lammoi.place(relx=0.2, rely=0.5, anchor=tkinter.CENTER)

        def callUpdateImageGUI():
            clear_frame()

            UpdateImageFrame = customtkinter.CTkFrame(rightSideFrame, height=500, width=600, fg_color="white")

            video_var = tk.StringVar(value="off")
            camera_var = tk.StringVar(value="off")

            def checkbox_event(var, other_var):
                if var.get() == "on":
                    other_var.set("off")

            # Create the first checkbox
            video_confirm = customtkinter.CTkCheckBox(UpdateImageFrame, text="Video",
                                                      command=lambda: checkbox_event(video_var, camera_var),
                                                      variable=video_var, onvalue="on", offvalue="off",
                                                      text_color="black")
            video_confirm.place(relx=0.1, rely=0.7)

            # Create the second checkbox
            camera_confirm = customtkinter.CTkCheckBox(UpdateImageFrame, text="Camera",
                                                       command=lambda: checkbox_event(camera_var, video_var),
                                                       variable=camera_var, onvalue="on", offvalue="off",
                                                       text_color="black")
            camera_confirm.place(relx=0.3, rely=0.7)


            def test_model():
                test_model.configure(state='disabled')
                if video_var.get() == "on":
                    print("Video")

                    def onclose():
                        window.configure(state="normal")
                        URL_detection.destroy()

                    def URL_detection_confirmation():
                        choice = URL_detection_entry.get()
                        formatted_choice = choice.replace("\\", "\\\\")
                        formatted_choice = choice.replace('"', '')
                        test_detection.testing_anti_model(formatted_choice, True)
                        onclose()

                    URL_detection = customtkinter.CTkToplevel()
                    window.configure(state="disabled")
                    URL_detection.geometry("500x200+200+100")
                    URL_option = StringVar()
                    URL_detection_entry = customtkinter.CTkEntry(URL_detection, textvariable=URL_option,
                                                                placeholder_text="URL", width=400)
                    URL_detection_entry.pack(pady=30, padx=50)
                    URL_vid_heed = customtkinter.CTkLabel(URL_detection,
                                                          text=r'Đường dẫn ví dụ: "C:\Users\ADMIN\OneDrive\Máy tính\quan(3).mp4"',
                                                          text_color="grey").place(rely=0.3, relx=0.1)
                    URL_detection_button = customtkinter.CTkButton(URL_detection, text="Confirm",
                                                                  command=URL_detection_confirmation).place(relx=0.35,
                                                                                                           rely=0.7)

                elif camera_var.get() == "on":
                    print("Camera")
                    test_detection.testing_anti_model(0,True)
                elif video_var.get() == "off" and camera_var.get() == "off":
                    messagebox.showwarning("Chưa chọn chế độ detect", "Vui lòng chọn chế độ detect")
                test_model.configure(state="normal")


            def activate_scanning_anti_sample():
                samples_name = scan_sample_entry.get()
                if samples_name:
                    scan.configure(state="disabled")
                    if video_var.get() == "on":
                        print("Video")

                        def onclose():
                            window.configure(state="normal")
                            URL_vid_path.destroy()

                        def URL_vid_path_confirmation():
                            choice = URL_vid_path_entry.get()
                            if choice:
                                formatted_choice = choice.replace("\\", "\\\\")
                                formatted_choice = choice.replace('"', '')
                                scanning.scanning(samples_name, formatted_choice)
                                onclose()
                            else: messagebox.showwarning("","Vui lòng điền URL")

                        URL_vid_path = customtkinter.CTkToplevel()
                        window.configure(state="disabled")
                        URL_vid_path.geometry("500x200+200+100")
                        URL_option = StringVar()
                        URL_vid_path_entry = customtkinter.CTkEntry(URL_vid_path, textvariable=URL_option,placeholder_text="URL", width=400)
                        URL_vid_path_entry.pack(pady=30, padx=50)
                        URL_vid_heed = customtkinter.CTkLabel(URL_vid_path,
                                                              text=r'Đường dẫn ví dụ: "C:\Users\ADMIN\OneDrive\Máy tính\quan(3).mp4"',text_color="grey").place(rely=0.3, relx=0.1)
                        URL_vid_path_button = customtkinter.CTkButton(URL_vid_path, text="Confirm",command=URL_vid_path_confirmation).place(relx=0.35, rely=0.7)

                    elif camera_var.get() == "on":
                        print("Camera")
                        scanning.scanning(samples_name, 0)
                    elif video_var.get() == "off" and camera_var.get() == "off":
                        messagebox.showwarning("Chưa chọn chế độ scan", "Vui lòng chọn chế độ Scan")
                    scan.configure(state="normal")
                else:
                    messagebox.showwarning("Warning", "Bạn chưa đặt tên folder trước khi lấy samples")

            def find_imgs_and_open_folder():
                encodeImg.configure(text="Encode Face File")
                # folder_name = "images"
                folder_name = "imgs_factory"
                for root, dirs, files in os.walk('.'):
                    if folder_name in dirs:
                        folder_path = os.path.join(root, folder_name)
                        os.startfile(folder_path)
                        break

            def open_training_result():
                folder_name = "training_log_softmax.csv"
                for root, dirs, files in os.walk('.'):
                    if folder_name in files:
                        folder_path = os.path.join(root, folder_name)
                        os.startfile(folder_path)
                        break
                    messagebox.showwarning("Không tồn tại", f'Không có kết quả training')
                    break


            imgs_folder = customtkinter.CTkButton(master=UpdateImageFrame,
                                                  text="Open Images Folder", width=80, text_color="white",
                                                  border_width=1,
                                                  border_color="black", corner_radius=20,
                                                  command=find_imgs_and_open_folder)
            imgs_folder.place(relx=0.1, rely=0.9)

            scan = customtkinter.CTkButton(master=UpdateImageFrame,
                                           text="Scan Anti Samples", width=80, text_color="white",
                                           border_width=1,
                                           border_color="black", corner_radius=20,
                                           command=activate_scanning_anti_sample)
            scan.place(relx=0.1, rely=0.8)

            scan_sample_entry = customtkinter.CTkEntry(master=UpdateImageFrame,
                                                       placeholder_text="Tên Folder", width=100, text_color="black",
                                                       border_width=2,
                                                       border_color="black", corner_radius=20, fg_color="white")
            scan_sample_entry.place(relx=0.1, rely=0.6)

            training_result = customtkinter.CTkButton(master=UpdateImageFrame,text='Result',text_color="white",width=50,border_width=1,corner_radius=20,border_color="black",command=open_training_result)
            training_result.place(relx=0.28, rely=0.6)

            test_model = customtkinter.CTkButton(master=UpdateImageFrame,text='Test Detection', text_color="white",width=70,corner_radius=20,border_width=1,border_color="black",command=test_model)
            test_model.place(relx=0.12,rely=0.25)

            heed_label= customtkinter.CTkLabel(UpdateImageFrame, text='Cắt ảnh 128x128', text_color="black").place(x=290,y=50)
            image_file = Image.open('img/face_example.png')
            Face_Example_image = customtkinter.CTkImage(light_image=image_file,size=(100,100))
            Face_Example_image_Label = customtkinter.CTkLabel(UpdateImageFrame,text='', image=Face_Example_image).place(x=400, y=10)

            temp1 = customtkinter.CTkLabel(UpdateImageFrame,text='Real',text_color="black").place(relx=0.55,rely=0.5)
            temp2 = customtkinter.CTkLabel(UpdateImageFrame, text='Fake', text_color="black").place(relx=0.75,rely=0.5)
            live_image_file = Image.open('img/live_example.png')
            Live_Example_image = customtkinter.CTkImage(light_image=live_image_file,size=(100,100))
            Live_Example_image_Label = customtkinter.CTkLabel(UpdateImageFrame,text='', image=Live_Example_image).place(relx=0.5,rely=0.6)
            fake_image_file = Image.open('img/fake_example.png')
            Fake_Example_image = customtkinter.CTkImage(dark_image=fake_image_file,size=(100,100))
            Fake_Example_image_Label = customtkinter.CTkLabel(UpdateImageFrame,text='', image=Fake_Example_image).place(relx=0.7,rely=0.6)

            def find_and_open_folder():
                encodeImg.configure(text="Encode Face File")
                folder_name = "images"
                # folder_name = "datasets"
                for root, dirs, files in os.walk('.'):
                    if folder_name in dirs:
                        folder_path = os.path.join(root, folder_name)
                        os.startfile(folder_path)
                        break

            def run_external_script1():
                os.system("face_recognition_model_face_detecting.py")
                # subprocess(["python","face_recognition_model_face_detecting.py"])
                # update the GUI
                encodeImg.configure(state='normal', text="Encoding Complete")
                window.update()

            def activate_encoding():
                encodeImg.configure(text="Encoding. . .")
                encodeImg.configure(state='disabled')
                Thread(target=run_external_script1).start()
                window.after(500)

            def find_anti_and_open_folder():
                anti_encodeImg.configure(text="Encode Live File")
                folder_name = "datasets-anti-spoofing"
                for root, dirs, files in os.walk('.'):
                    if folder_name in dirs:
                        folder_path = os.path.join(root, folder_name)
                        os.startfile(folder_path)
                        break

            def run_external_script2():
                os.system("cnn_model_anti_detecting.py")
                # update the GUI
                anti_encodeImg.configure(state='normal', text="Encoding Complete")
                window.update()

            def activate_anti_encoding():
                anti_encodeImg.configure(text="Encoding. . .")
                anti_encodeImg.configure(state='disabled')
                Thread(target=run_external_script2).start()
                window.after(500)

            face_folder = customtkinter.CTkButton(master=UpdateImageFrame,
                                                  text="Open Face Folder", width=150, text_color="white", border_width=1,
                                                  border_color="black", corner_radius=20, command=find_and_open_folder)
            face_folder.place(relx=0.1, rely=0.01)
            encodeImg = customtkinter.CTkButton(master=UpdateImageFrame,
                                                text="Encode Face File", width=150, text_color="white", border_width=1,
                                                border_color="black", corner_radius=20, command=activate_encoding)
            encodeImg.place(relx=0.1, rely=0.1)

            anti_live_folder = customtkinter.CTkButton(master=UpdateImageFrame,
                                                       text="Open Live Folder", width=150, text_color="white",
                                                       border_width=1,
                                                       border_color="black", corner_radius=20,
                                                       command=find_anti_and_open_folder)
            anti_live_folder.place(relx=0.1, rely=0.4)
            anti_encodeImg = customtkinter.CTkButton(master=UpdateImageFrame,
                                                     text="Encode Live File", width=150, text_color="white",
                                                     border_width=1,
                                                     border_color="black", corner_radius=20,
                                                     command=activate_anti_encoding)
            anti_encodeImg.place(relx=0.1, rely=0.493)
            UpdateImageFrame.pack(fill=tkinter.BOTH, padx=10, pady=10)


        def callUpdateScoreGUI():
            clear_frame()
            UpdateScoreFrame = customtkinter.CTkFrame(rightSideFrame, height=500, width=600, fg_color="white")
            idSinhVien = customtkinter.CTkEntry(master=UpdateScoreFrame,
                                                placeholder_text="Nhập ID Sinh viên cần update điểm", width=400,
                                                height=25, )

            def updateScore():
                id = idSinhVien.get()
                mycursor = connection.mydb.cursor()
                sql = "SELECT idSinhVien FROM SinhVien"

                mycursor.execute(sql)

                myresult = mycursor.fetchall()

                if id == '':
                    messagebox.showinfo("Chưa nhập ID", "Vui lòng nhập ID và thử lại")
                    return
                else:
                        try:
                            id = int(id)
                            found = False
                            for tup in myresult:
                                if id in tup:
                                    found = True
                                    break
                            if found:
                                sql_detail_score = """select mh.mamonhoc,mh.tenmon, ct.diemgiuaki, ct.diemcuoiki
                                                        from BangDiemChiTiet as ct
                                                        join MonHoc as mh
                                                        where ct.bangdiemchitiet = (
                                                                select tt.bangdiemchitiet
                                                                from ThanhTich as tt
                                                                where tt.mabangdiem = (select sv.mabangdiem
                                                                                        from SinhVien as sv
                                                                                        where sv.idSinhVien = %s)
                                                                ) and mh.mamonhoc = ct.mamonhoc;
                                                        """
                                data = (id,)
                                mycursor.execute(sql_detail_score, data)
                                result = mycursor.fetchall()

                                detailScoreTop = customtkinter.CTkToplevel(UpdateScoreFrame)
                                detailScoreTop.title("Điểm chi tiết")
                                detailScoreTop.geometry("700x400")
                                tree = ttk.Treeview(detailScoreTop, columns=("mamon", "tenmon", "diemgiuaki", "diemcuoiki"),
                                                    show="headings", height=17)
                                tree.heading("mamon", text="Mã Môn Học")
                                tree.heading("tenmon", text="Tên Môn Học")
                                tree.heading("diemgiuaki", text="Điểm Giữa Kì")
                                tree.heading("diemcuoiki", text="Điểm Cuối Kì")
                                for x in result:
                                    tree.insert('', "end", iid=x[0], values=x)
                                tree.pack(padx=10, pady=10)

                                def updateScoreSV():
                                    def edit():
                                        selected = tree.selection()[0]
                                        diemgiuaki = entryGiuaKi.get()
                                        diemcuoiki = entryCuoiKi.get()
                                        sql = """UPDATE BangDiemChiTiet
                                                SET diemgiuaki = %s, diemcuoiki = %s
                                                WHERE bangdiemchitiet IN (
                                                    SELECT bangdiemchitiet
                                                    FROM (
                                                        SELECT DISTINCT tt.bangdiemchitiet
                                                        FROM SinhVien AS sv
                                                        JOIN ThanhTich AS tt ON sv.mabangdiem = tt.mabangdiem
                                                        JOIN BangDiemChiTiet AS ct ON tt.bangdiemchitiet = ct.bangdiemchitiet
                                                        WHERE sv.idSinhVien = %s
                                                    ) AS subquery
                                                )
                                                AND mamonhoc = %s;
                                                """
                                        data = (float(diemgiuaki), float(diemcuoiki), id, selected)
                                        try:
                                            mycursor.execute(sql, data)
                                            connection.mydb.commit()
                                            editFrame.destroy()
                                            # Fetch the updated scores
                                            mycursor.execute(sql_detail_score, (id,))
                                            updated_result = mycursor.fetchall()

                                            # Clear the existing tree
                                            for item in tree.get_children():
                                                tree.delete(item)

                                            # Repopulate the tree with the updated scores
                                            for x in updated_result:
                                                tree.insert('', "end", iid=x[0], values=x)
                                        except Exception as e:
                                            print("Error updating database:", e)
                                            messagebox.showerror("Database Error", "Failed to update the database.")

                                    if idSinhVien.winfo_exists():
                                        idSinhVien.delete(0, tkinter.END)
                                    editFrame = customtkinter.CTkToplevel(height=400, width=650, fg_color="white")
                                    editFrame.title("Sửa điểm")
                                    entryGiuaKi = customtkinter.CTkEntry(editFrame, placeholder_text="Điểm giữa kì", width=300,
                                                                         height=45)
                                    entryCuoiKi = customtkinter.CTkEntry(editFrame, placeholder_text="Điểm cuối kì", width=300,
                                                                         height=45)
                                    editBtn = customtkinter.CTkButton(master=editFrame, text="OK", width=60, height=40,
                                                                      command=edit)
                                    editBtn.place(relx=0.4, rely=0.6)
                                    entryGiuaKi.place(relx=0.2, rely=0.2)
                                    entryCuoiKi.place(relx=0.2, rely=0.4)

                                editBtn = customtkinter.CTkButton(detailScoreTop, width=80, height=32, border_width=0,
                                                                  corner_radius=8, text="Sửa", command=updateScoreSV)
                                editBtn.place(relx=0.5, rely=0.8, anchor=tkinter.CENTER)
                                # detailScoreTop.mainloop()
                                print(result)

                            else:
                                messagebox.showinfo("Không có sinh viên này", "Không có hình ảnh tương ứng với ID này")
                        except Exception as e:
                            print(e)
                            messagebox.showerror(title="Lỗi dữ liệu nhập vào",message="ID Không hợp lệ, phải là số")

                idSinhVien.delete(0, tkinter.END)

            idSinhVien.place(relx=0.2, rely=0.2, )
            editBtn = customtkinter.CTkButton(master=UpdateScoreFrame,
                                              width=80,
                                              height=32,
                                              border_width=0,
                                              corner_radius=8,
                                              text="Sửa",
                                              command=updateScore
                                              )
            editBtn.place(relx=0.5, rely=0.8, anchor=tkinter.CENTER)

            UpdateScoreFrame.pack(fill=tkinter.BOTH, padx=10, pady=10)

        window = customtkinter.CTk()
        window.geometry('800x400+350+120')
        window.title("Admin")
        window.config(background="#C7C7C7")

        headerFrame = customtkinter.CTkFrame(window, width=600, height=50, fg_color="white",border_width=1)
        adminImage = Image.open("img/admin.png")
        label1 = customtkinter.CTkLabel(master=headerFrame, width=40, height=80, text="",
                                        image=customtkinter.CTkImage(adminImage))
        label1.place(relx=0.45, rely=0.5, anchor=tkinter.CENTER)
        label2 = customtkinter.CTkLabel(master=headerFrame, text="Admin")
        label2.place(relx=0.5, rely=0.5, anchor=tkinter.CENTER)
        headerFrame.pack(side=tkinter.TOP, fill=tkinter.X)

        # Left Side
        leftSideFrame = customtkinter.CTkFrame(master=window, height=50, width=160, fg_color="white",border_width=1)
        dashboardBtn = customtkinter.CTkButton(master=leftSideFrame, text="Dashboard", anchor="center", fg_color="#60c9db",
                                               text_color="#000", command=callDashboardGUI)
        dashboardBtn.place(x=10, y=20)
        studentBtn = customtkinter.CTkButton(master=leftSideFrame, text="Student", anchor="center", fg_color="#60c9db",
                                             text_color="#000", command=callStudentGUI)
        studentBtn.place(x=10, y=60)

        courseBtn = customtkinter.CTkButton(master=leftSideFrame, text="Course", anchor="center", fg_color="#60c9db",
                                            text_color="#000", command=callCourseGUI)
        courseBtn.place(x=10, y=100)

        updateCLassBtn = customtkinter.CTkButton(master=leftSideFrame, text="Class", anchor="center",
                                                 fg_color="#60c9db",
                                                 text_color="#000", command=callUpdateClass)
        updateCLassBtn.place(x=10, y=140)

        updateCoVanBtn = customtkinter.CTkButton(master=leftSideFrame, text="Adviser", anchor="center", fg_color="#60c9db",
                                            text_color="#000", command=callUpdateCoVan)
        updateCoVanBtn.place(x=10, y=180)

        updateImageBtn = customtkinter.CTkButton(master=leftSideFrame, text="Update Image", anchor="center",
                                                 fg_color="#60c9db", text_color="#000", command=callUpdateImageGUI)
        updateImageBtn.place(x=10, y=220)

        updateScoreBtn = customtkinter.CTkButton(master=leftSideFrame, text="Update Score", anchor="center",
                                                 fg_color="#60c9db", text_color="#000", command=callUpdateScoreGUI)
        updateScoreBtn.place(x=10, y=260)

        leftSideFrame.pack(side=tkinter.LEFT, fill=tkinter.Y, pady=10)

        # Right Side

        rightSideFrame = customtkinter.CTkFrame(window, height=300, width=600, fg_color="black",border_width=1)
        intro = customtkinter.CTkLabel(master=rightSideFrame,
                                       text="Chào mừng admin trở lại với trang quản lý thông tin của sinh viên !!!",
                                       font=('Times New Roman', 24, 'bold'),
                                       width=400, wraplength=400)
        intro.place(relx=0.2, rely=0.4)
        rightSideFrame.pack(side=tkinter.RIGHT, fill=tkinter.BOTH, padx=10, pady=10)
        window.protocol("WM_DELETE_WINDOW", on_close)
        window.resizable(False, False)
        window.mainloop()

    # print(idMaBangDiem)
    # print(idCoVan)
    # print(id)

    if admin_window and username == 'admin' and password == 'admin':
        value = admin()
        admin_window = None

    if value is None:
        # print("Value is None")
        pass
    else:
        # print("Value is True")
        main()
        value = None

    if verify:
        main()
    else:
        LOGIN()


